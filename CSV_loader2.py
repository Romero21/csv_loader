
import time, os, os.path, glob, shutil, re
from subprocess import check_output
from pywinauto.application import Application
from datetime import datetime
from threading import Thread
import pandas as pd, csv
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.styles import Alignment

space = " "
print('********************************************************************')
print(f'*{66*space}*')
print(f'*{4*space}**********************************************************{4*space}*')
print(f'*{4*space}********************CSV LOADER V1.02**********************{4*space}*')
print(f'*{4*space}****************Created by Roman Smycek*******************{4*space}*')
print(f'*{4*space}**********************************************************{4*space}*')
print(f'*{4*space}**********************************************************{4*space}*')
print(f'*{66*space}*')
print('********************************************************************')

directory = os.getcwd()
file_type = '\*csv'
folder = '\logs'
subfolder = '\logs\\recent_files'


def Get_latest_CSV():
    global new_name_CSV
    global latest_CSV
    latest_CSV = max(glob.glob(new_path + file_type), key=os.path.getctime)
    new_name_CSV = os.path.join(new_path, "device.csv")
    os.rename(latest_CSV, new_name_CSV)
    latest_CSV = r"C:\Us" + new_name_CSV

def translator():
        try:
            translator = r'\ReadElmDeviceLogEx.exe /p '
            cmd = r'CMD /c '
            print(cmd + directory + translator + new_path)
            app = Application().start(cmd + directory + translator + new_path, create_new_console=True, wait_for_idle=False, timeout=20)
            main_dlg = app.window(title=r'C:\Windows\system32\cmd.exe')
            main_dlg.wait('visible')
        except:
            pass


def unzip():
        print("Yes, this is a zip file")
        created_path = glob.glob(new_path + types)
        new_dir_files = str(created_path)
        new_dir_files = max(created_path, key=os.path.getctime)
        shutil.unpack_archive(new_dir_files, new_path)
            

def GetFileExtension():
    global types
    global new_path
    global latest_File
    types = '\*'
    log_dir = (directory + folder)
    new_log_dir = (directory + subfolder)
    new_path = os.path.join(new_log_dir, datetime.now().strftime("%d_%m_%Y-%H_%M_%S"))
    try:
        os.mkdir(new_path)
    except:
        print("Path is already done.")
    all_files = os.listdir(log_dir)
    for files in all_files:
        if files.endswith('.log'):
            shutil.move(os.path.join(log_dir,files), os.path.join(new_path,files))
            translator()
            csv_exists()
        if files.endswith('.csv'):
            shutil.move(os.path.join(log_dir,files), os.path.join(new_path,files))
            csv_exists()
        if files.endswith('.zip'):
            shutil.move(os.path.join(log_dir,files), os.path.join(new_path,files))
            unzip()
            translator()
            csv_exists()
        if files.endswith('.zip' or '.log' or '.csv') is False:
            print(f"\nFound files:\t{files}\n")
            print('No file or unsupported format found. \nPlease insert supported file into LOGS directory.')
            os.system('pause')
            GetFileExtension()


def csv_to_excel():
    Get_latest_CSV()
    path = new_path + '\device.csv'
    path2 = new_path + '\device.csvencode.csv'
    date = datetime.now().strftime("%M_%S")
    final_file = new_path + f'\device{date}.xlsx'

    with open(path, 'r', encoding='utf-16-le', errors='ignore') as infile, open(path2, 'w') as outfile:
         inputs = csv.reader(infile)
         output = csv.writer(outfile)
         for row in inputs:
             output.writerow(row)

    df = pd.read_csv(path2)
    df.to_excel(final_file)

    wb = load_workbook(filename = final_file)
    ws = wb['Sheet1']

    patFill_1 = PatternFill(patternType='solid', fgColor= 'b1ff94')
    patFill_2 = PatternFill(patternType='solid', fgColor= 'e2ff94')
    patFill_3 = PatternFill(patternType='solid', fgColor= 'ffe094')
    rowPatFill = PatternFill(patternType='solid', fgColor= 'd5ffbd')


    for cell in ws["H:H"]:
        cell.fill = patFill_1
    for cell in ws["K:K"]:
        cell.fill = patFill_2
    for cell in ws["L:L"]:
        cell.fill = patFill_3

    def fillRows(count):
        while count < 300:
            for row in ws[(count):(count)]:
                row.fill = rowPatFill
            count += 2

    fillRows(1)


    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 25
    ws.column_dimensions["C"].width = 25
    ws.column_dimensions["D"].width = 40
    ws.column_dimensions["E"].width = 10
    ws.column_dimensions["F"].width = 30


    for column in ws["K:K"]:
        column.alignment = Alignment(horizontal='left')
    ws.delete_cols(9, 2)
    ws.delete_cols(1, 4)
    ws.delete_cols(7, 1)
    wb.save(final_file)

    check_output(['start', final_file], shell=True)


def csv_exists():
    find = re.findall('csv', str(os.listdir(new_path)))
    if find:
        print("CSV Loading...")
        csv_to_excel()

    else:
        time.sleep(0.25)
        csv_exists()

if __name__ == '__main__':
    Thread(target=GetFileExtension()).start()

    




                             
